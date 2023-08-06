"""Write to Excel.

We use an ExcelExport class that does a lot of the heavy lifting, formatting-wise.

This means that fields will get proper styling, number formatting, data validation
(dropdown selections), and alignment.

There are some named styles you can use to change cell appearance, apply them by name:

- "title"       blue5 bg, 16pt bold
- "key"         lightblue3 bg
- "name"        lightblue3 bg
- "label"       blue5 bg, bold
- "error"       lightred3 bg
- "warning"     lightorange3 bg
- "changed"     lightgreen3 bg
- "deleted"     lightred5 bg, strikethrough
- "disabled"    lightgray2 bg
- "muted"       lightgray4 bg


Simple usage:

```
excel = ExcelExport("Unicat missing images")
cell = excel.headercell(row=1, column=1, value="gid")
cell = excel.fieldcell(
    row=row,
    column=column,
    value=fieldvalue,
    field=field,
    language=language,
)
cell.style = "warning"
```

Header cells automatically get the "name" style.
Alignment and wrapping, and number formatting is based on field settings.
Field values are converted based on type, e.g. lists will be stored as newline-separated
text.

Override anything on the returned cell using standard `openpyxl` techniques.
"""
from openpyxl import DEFUSEDXML, Workbook

if not DEFUSEDXML:
    print("missing defusedxml!")
    import sys

    sys.exit(1)


from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

COMMENT_AUTHOR = "Unicat"

from .styles import (
    HEADERFIELDTYPE_ALIGNMENT,
    FIELDTYPE_ALIGNMENT,
    FIELDTYPE_NUMBER_FORMAT,
    FIELDTYPE_WIDTHS,
    add_workbook_styles,
)

from ..utils import convert_fielddata_to_value


def is_empty_cell(cell):
    return not cell or not cell.value or not len(cell.value)


def is_deleted(worksheet, cell):
    return cell.font.strike == True


def cell_comment(cell, comment):
    cell.comment = Comment(comment, COMMENT_AUTHOR)
    cell.comment.width = 300
    cell.comment.height = 150


class ExcelExport:
    def __init__(self, worksheetname):
        self.wb = Workbook()
        add_workbook_styles(self.wb)
        self.ws = self.wb.active
        self.ws.title = worksheetname
        self._validations = {}

    def headercell(
        self, row, column, value, field=None, width=None, style=None, comment=None
    ):
        cell = self.ws.cell(column=column, row=row)
        cell.value = value
        cell.style = "name"
        if field:
            cell.alignment = HEADERFIELDTYPE_ALIGNMENT[field.type]
            self.ws.column_dimensions[
                get_column_letter(column)
            ].width = FIELDTYPE_WIDTHS[field.type]
        if width is not None:
            self.ws.column_dimensions[get_column_letter(column)].width = width
        if style is not None:
            cell.style = style
        if comment is not None:
            cell.comment = Comment(comment, COMMENT_AUTHOR)
        return cell

    def fieldcell(
        self,
        row,
        column,
        value,
        field,
        width=None,
        language=None,
        style=None,
        comment=None,
    ):
        cell = self.ws.cell(column=column, row=row)
        cell.value = convert_fielddata_to_value(field.type, value)
        cell.alignment = FIELDTYPE_ALIGNMENT[field.type]
        self.ws.column_dimensions[get_column_letter(column)].width = FIELDTYPE_WIDTHS[
            field.type
        ]
        number_format = FIELDTYPE_NUMBER_FORMAT[field.type]
        if (
            field.type == "decimal"
            and "decimals" in field.options
            and field.options["decimals"]
        ):
            number_format = FIELDTYPE_NUMBER_FORMAT["decimalN"] + (
                "0" * int(field.options["decimals"])
            )
        cell.number_format = number_format
        # add validation -> dropdown select boxes
        validationkey = field.gid + "." + language
        if validationkey not in self._validations:
            if (
                field.type in ("textline", "textlist")
                and "values" in field.options
                and field.options["values"]
            ):
                values = ",".join(
                    [
                        value[language]
                        for value in field.options["values"]
                        if "," not in value[language]
                    ]
                )
                if len(values) < 255:
                    dv = DataValidation(
                        type="list",
                        formula1=f'"{values}"',
                        errorStyle="information",
                        errorTitle="Value not in list",
                        error="It can be saved regardless if this is intentional.",
                        allow_blank=True,
                    )
                    self.ws.add_data_validation(dv)
                    self._validations[validationkey] = dv
        if validationkey in self._validations:
            dv = self._validations[validationkey]
            dv.add(cell)
        if width is not None:
            self.ws.column_dimensions[get_column_letter(column)].width = width
        if style is not None:
            cell.style = style
        if comment is not None:
            cell.comment = Comment(comment, COMMENT_AUTHOR)
        return cell

    def save(self, filepath):
        for index, _ in enumerate(self.ws.rows, 1):
            self.ws.row_dimensions[index].bestFit = True
        self.wb.save(filepath)
