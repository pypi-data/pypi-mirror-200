from openpyxl.styles import (
    NamedStyle,
    PatternFill,
    Font,
    Alignment,
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="top")
ALIGN_CENTER = Alignment(horizontal="center", vertical="top")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="top")
ALIGN_GENERAL = Alignment(horizontal="general", vertical="top")
WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrapText=True)
WRAP_CENTER = Alignment(horizontal="center", vertical="top", wrapText=True)
WRAP_RIGHT = Alignment(horizontal="right", vertical="top", wrapText=True)
WRAP_GENERAL = Alignment(horizontal="general", vertical="top", wrapText=True)

FILL_NONE = PatternFill(fill_type=None)
FILL_TITLE = PatternFill("solid", fgColor="c9e0f2")  # $cc-blue5
FILL_KEY = PatternFill("solid", fgColor="e4eff8")  # $cc-lightblue3
FILL_NAME = PatternFill("solid", fgColor="e4eff8")  # $cc-lightblue3
FILL_LABEL = PatternFill("solid", fgColor="c9e0f2")  # $cc-blue5
FILL_ERROR = PatternFill("solid", fgColor="f9e2e5")  # $cc-lightred3
FILL_WARNING = PatternFill("solid", fgColor="fdf3de")  # $cc-lightorange3
FILL_CHANGED = PatternFill("solid", fgColor="eff7e2")  # $cc-lightgreen3
FILL_DELETED = PatternFill("solid", fgColor="fff5f6")  # $cc-lightred5
FILL_DISABLED = PatternFill("solid", fgColor="dfdfe1")  # $cc-lightgray2
FILL_MUTED = PatternFill("solid", fgColor="f4f4f5")  # $cc-lightgray4

FONT_DELETED = Font(strike=True, color="cd4942")  # $cc-red2

NAMED_STYLES = {}
NAMED_STYLES["title"] = NamedStyle(
    name="title", fill=FILL_TITLE, font=Font(size=16, bold=True), alignment=ALIGN_LEFT
)
NAMED_STYLES["key"] = NamedStyle(name="key", fill=FILL_KEY, alignment=WRAP_LEFT)
NAMED_STYLES["name"] = NamedStyle(name="name", fill=FILL_NAME, alignment=WRAP_LEFT)
NAMED_STYLES["value"] = NamedStyle(name="value", alignment=WRAP_LEFT)
NAMED_STYLES["label"] = NamedStyle(
    name="label", fill=FILL_LABEL, font=Font(bold=True), alignment=WRAP_LEFT
)
NAMED_STYLES["error"] = NamedStyle(name="error", fill=FILL_ERROR)
NAMED_STYLES["warning"] = NamedStyle(name="warning", fill=FILL_WARNING)
NAMED_STYLES["changed"] = NamedStyle(name="changed", fill=FILL_CHANGED)
NAMED_STYLES["deleted"] = NamedStyle(
    name="deleted", fill=FILL_DELETED, font=FONT_DELETED
)
NAMED_STYLES["disabled"] = NamedStyle(name="disabled", fill=FILL_DISABLED)
NAMED_STYLES["muted"] = NamedStyle(name="muted", fill=FILL_MUTED)

FIELDTYPE_ALIGNMENT = {
    "text": WRAP_LEFT,
    "textline": ALIGN_LEFT,
    "textlist": WRAP_LEFT,
    "number": ALIGN_RIGHT,
    "decimal": ALIGN_RIGHT,
    "boolean": ALIGN_CENTER,
    "code": WRAP_LEFT,
    "barcode": ALIGN_LEFT,
    "class": WRAP_LEFT,
    "classlist": WRAP_LEFT,
    "image": ALIGN_LEFT,
    "imagelist": WRAP_LEFT,
    "file": ALIGN_LEFT,
    "filelist": WRAP_LEFT,
    "record": ALIGN_LEFT,
    "recordlist": WRAP_LEFT,
    "fieldpicker": WRAP_LEFT,
}

HEADERFIELDTYPE_ALIGNMENT = {
    "text": WRAP_LEFT,
    "textline": WRAP_LEFT,
    "textlist": WRAP_LEFT,
    "number": WRAP_RIGHT,
    "decimal": WRAP_RIGHT,
    "boolean": WRAP_CENTER,
    "code": WRAP_LEFT,
    "barcode": WRAP_LEFT,
    "class": WRAP_LEFT,
    "classlist": WRAP_LEFT,
    "image": WRAP_LEFT,
    "imagelist": WRAP_LEFT,
    "file": WRAP_LEFT,
    "filelist": WRAP_LEFT,
    "record": WRAP_LEFT,
    "recordlist": WRAP_LEFT,
    "fieldpicker": WRAP_LEFT,
}

FIELDTYPE_NUMBER_FORMAT = {
    "text": "@",  # text
    "textline": "@",
    "textlist": "@",
    "number": "#,##0",
    "decimal": "#,##0",
    "decimalN": "#,##0.",
    "boolean": "@",
    "code": "@",
    "barcode": "@",
    "class": "@",
    "classlist": "@",
    "image": "@",
    "imagelist": "@",
    "file": "@",
    "filelist": "@",
    "record": "@",
    "recordlist": "@",
    "fieldpicker": "@",
}

FIELDTYPE_WIDTHS = {
    "text": 50,
    "textline": 18,
    "textlist": 18,
    "number": 12,
    "decimal": 12,
    "decimalN": 12,
    "boolean": 12,
    "code": 50,
    "barcode": 18,
    "class": 12,
    "classlist": 36,
    "image": 36,
    "imagelist": 36,
    "file": 36,
    "filelist": 36,
    "record": 36,
    "recordlist": 36,
    "fieldpicker": 36,
}


def add_workbook_styles(workbook):
    for named_style in NAMED_STYLES.values():
        try:
            workbook.add_named_style(named_style)
        except ValueError:
            pass  # already exists


def reset_worksheet_styles(worksheet):
    """Reset styles"""
    for row in worksheet.rows:
        for cell in row:
            cell.fill = FILL_NONE
            cell.comment = None
