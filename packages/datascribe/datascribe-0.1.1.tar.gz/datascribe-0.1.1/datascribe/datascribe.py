import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

def save_dataset_notes(df, title, reason):
    """
    This function takes in a DataFrame, extracts details of the dataset, and saves them to an Excel sheet.
    """
    # Get the dataset attributes
    shape = df.shape
    url = ""
    date = datetime.now().strftime("%Y-%m-%d")
    time = datetime.now().strftime("%H:%M:%S")

    # Create a new DataFrame to hold the dataset details
    data = {'Shape': shape,
            'URL': url,
            'Title': title,
            'Reason': reason,
            'Date': date,
            'Time': time}
    df_details = pd.DataFrame(data, index=[0])

    # Load the existing workbook and sheet
    try:
        wb = load_workbook('dataset_notes.xlsx')
    except FileNotFoundError:
        # If the file doesn't exist, create a new one with the headers
        headers = ['Shape', 'URL', 'Title', 'Reason', 'Date', 'Time']
        df_headers = pd.DataFrame(columns=headers)
        df_headers.to_excel('dataset_notes.xlsx', index=False)
        wb = load_workbook('dataset_notes.xlsx')
    sheet = wb.active
    
    # Append the new data to the next available row
    next_row = sheet.max_row + 1
    for r in dataframe_to_rows(df_details, index=False, header=False):
        for idx, val in enumerate(r, start=1):
            sheet.cell(row=next_row, column=idx, value=val)
        next_row += 1
    
    # Save the workbook
    wb.save("dataset_notes.xlsx")
