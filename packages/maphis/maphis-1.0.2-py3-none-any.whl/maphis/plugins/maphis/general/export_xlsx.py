from pathlib import Path
import typing
from typing import Optional

import openpyxl
import openpyxl.worksheet.worksheet as worksheet
from openpyxl.styles import Color, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule, Rule
from PySide6.QtWidgets import QApplication

from maphis.common.utils import attempt_to_save_with_retries
from maphis.common.common import Info
from maphis.common.plugin import GeneralAction, ActionContext
from maphis.common.state import State
from maphis.plugins.maphis.general.common import _ndarray_export_routine, _filter_by_NDArray, \
    _group_measurements_by_sheet, _tabular_export_routine, show_export_success_message, get_prop_tuple_list, StyledCells


class ExportXLSX(GeneralAction):
    """
    GROUP: export
    NAME: Export to xlsx
    KEY: export_xlsx
    DESCRIPTION: Exports measurements in XLSX format.
    """

    def __init__(self, info: Optional[Info] = None):
        super().__init__(info)

    def __call__(self, state: State, context: ActionContext) -> None:
        wb = openpyxl.Workbook()
        prop_tuple_list = get_prop_tuple_list(context)
        # self._tabular_export_routine(ws.append)
        nd_props, other_props = _filter_by_NDArray(prop_tuple_list, context)

        sheet_grouped_props = _group_measurements_by_sheet(other_props, context)

        for sheet_name, prop_list in sheet_grouped_props.items():
            ws: worksheet.Worksheet = wb.create_sheet(sheet_name)
            styles = _tabular_export_routine(prop_list, ws.append, context)
            for style in styles:
                for cell in style.cells:
                    ws.cell(*cell).style = style.style

        sheet_nd_props = _group_measurements_by_sheet(nd_props, context)

        for sheet_name, prop_list in sheet_nd_props.items():
            ws = wb.create_sheet(sheet_name)
            cell_styles: typing.List[StyledCells] = _ndarray_export_routine(prop_list, ws.append, context)
            for styled_cells in cell_styles:
                for cell in styled_cells.cells:
                    ws.cell(*cell).style = styled_cells.style

        if 'Sheet' in wb:
            wb.remove(wb['Sheet'])

        # wb.save(str(context.storage.location / f'{context.current_label_name}_results.xlsx'))

        final_path, _ = attempt_to_save_with_retries(_save, 'Save as...', 'single_file', ['xlsx'], QApplication.activeWindow(),
                                     path=context.storage.location / f'{context.current_label_name}_results.xlsx',
                                     object=wb)
        if final_path is not None:
            show_export_success_message(final_path.parent, [final_path.name], context)

        # show_export_success_message(context.storage.location, [f'{context.current_label_name}_results.xlsx'], context)


    # @property
    # def group(self) -> str:
    #     return 'export'


def _save(path: Path, object: openpyxl.Workbook):
    object.save(path)
